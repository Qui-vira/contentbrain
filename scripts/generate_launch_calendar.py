"""Generate 30-Day Launch Calendar XLSX and Launch To-Do Tracker XLSX."""
import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', '06-Drafts')
CAL_PATH = os.path.join(OUT_DIR, '30-day-launch-calendar.xlsx')
TODO_PATH = os.path.join(OUT_DIR, 'launch-todo.xlsx')

# ── PALETTE ───────────────────────────────────────────────────────────────────
C = {
    'dark':     '1A1A2E',
    'red':      'E94560',
    'gray':     '444444',
    'light':    'F5F5F5',
    'white':    'FFFFFF',
    'mid':      'CCCCCC',
    'green':    '2D6A4F',
    'blue':     '1D4E89',
    'gold':     'B5862A',
    'orange':   'E07B39',
    'purple':   '6B3FA0',
    'teal':     '1A6B6B',
    'rose':     'C0392B',
    'posted':   'D5F0D5',
    'draft':    'FFF5D5',
    'locked':   'E8E8E8',
}

def fill(hex_col):
    return PatternFill('solid', fgColor=hex_col)

def font(hex_col='000000', bold=False, size=9, name='Calibri'):
    return Font(color=hex_col, bold=bold, size=size, name=name)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style='medium', color='1A1A2E')
    return Border(left=s, right=s, top=s, bottom=s)

def align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(cell, text, bg=C['dark'], fg=C['white'], bold=True, size=9):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = font(fg, bold, size)
    cell.alignment = align('center', 'center', False)
    cell.border = border_thin()

def style_cell(cell, value, bg=C['white'], fg=C['gray'], bold=False,
               h='left', v='top', wrap=True):
    cell.value = value
    cell.fill = fill(bg)
    cell.font = font(fg, bold)
    cell.alignment = align(h, v, wrap)
    cell.border = border_thin()

# ── WEEK COLORS ───────────────────────────────────────────────────────────────
WEEK_BG = {
    1: 'EAF4FF',  # light blue — reconnection
    2: 'FFF5E6',  # light amber — authority + teaser
    3: 'EAF5EA',  # light green — announcement
    4: 'FFF0F0',  # light rose — promotion
    5: 'F0E6FF',  # light purple — final push
}

PLATFORM_COLOR = {
    'X/Twitter':  C['dark'],
    'LinkedIn':   C['blue'],
    'TikTok':     C['teal'],
    'Instagram':  C['rose'],
    'All':        C['purple'],
    'Telegram':   C['green'],
}

CATEGORY_COLOR = {
    'Personality':        'FFE5EC',
    'Value':              'E5F5FF',
    'Authority':          'FFF5CC',
    'Freebie':            'E5FFE5',
    'Transformational':   'F5E5FF',
    'Lead Gen':           'FFE8CC',
    'Weekly Series':      'CCF5FF',
    'Money Proof':        'FFFACC',
    'Course Launch':      'FFCCCC',
    'Engagement':         'EEEEEE',
}

STATUS_COLOR = {
    'POSTED':     C['posted'],
    'LOCKED':     C['locked'],
    'DRAFT':      C['draft'],
    'SCHEDULED':  'E0F0FF',
}

# ── CALENDAR DATA ─────────────────────────────────────────────────────────────
# Columns: day, date, week, slot, platform, segment, category, hook, goal, format, cta, monetize, series, status, notes

START = date(2026, 4, 3)

POSTS = [
    # ── WEEK 1: RECONNECTION (Days 1–7) ──────────────────────────────────────

    # Day 1 — LOCKED (already posted)
    (1, 'Morning', 'X/Twitter', 'All', 'Personality',
     '"I disappeared. Not because I gave up..." — Comeback thread (5 tweets). CTA: "What did I miss?"',
     'Community', 'Thread', 'Reply engagement', 'N', '', 'LOCKED',
     'Day 1 locked. Already posted. Baseline for all future content.'),
    (1, 'Afternoon', 'X/Twitter', 'All', 'Personality',
     '"Most creators disappear because they burned out. I disappeared because I stopped trusting the process..."',
     'Reach', 'Tweet', '', 'N', '', 'LOCKED', 'Day 1 locked.'),
    (1, 'Morning', 'LinkedIn', 'All', 'Personality',
     '"I went silent for months. Not because the business failed..." — Long-form founder story',
     'Authority', 'Long-form post', 'Comments', 'N', '', 'LOCKED', 'Day 1 locked.'),
    (1, 'Afternoon', 'Instagram', 'All', 'Personality',
     'Cinematic reel (34s) with subtitles — "What I Built In Silence"',
     'Reach', 'Reel', 'DM CTA', 'N', '', 'LOCKED', 'Day 1 locked. Already posted via Graph API.'),
    (1, 'Evening', 'TikTok', 'All', 'Personality',
     'Talking head BTS (30s): "I disappeared for months. Here\'s why."',
     'Reach', 'TikTok video', 'Follow CTA', 'N', '', 'LOCKED', 'Day 1 locked. 1 TikTok posted.'),

    # Day 2 — Loss story → system
    (2, 'Morning', 'X/Twitter', 'All', 'Personality',
     '"I lost $4,000 in a single trade last year because I ignored my own rules. Then I built a system that can\'t ignore them."',
     'Reach', 'Thread (3 tweets)', 'Engagement', 'N', '', 'DRAFT',
     'Loss → lesson → system. Vulnerability + expertise formula.'),
    (2, 'Evening', 'X/Twitter', 'All', 'Authority',
     '"Unpopular opinion: 90% of trading signals on CT are just gambling disguised as analysis."',
     'Reach', 'Tweet', 'Reply bait', 'N', '', 'DRAFT', 'Bold take. Provokes replies.'),
    (2, 'Morning', 'LinkedIn', 'All', 'Authority',
     '"I lost $4,000 on a single trade. Then I did what most traders never do — turned the loss into an engineering problem."',
     'Authority', 'Long-form post', 'DM SYSTEM', 'N', '', 'DRAFT', 'Professional reframe of X thread.'),
    (2, 'Morning', 'Instagram', 'All', 'Personality',
     '"I lost $4,000 in one trade." — Carousel (7 slides): mistake, emotion, realization, system, result.',
     'Reach', 'Carousel', 'DM SYSTEM', 'N', '', 'DRAFT', 'Dark bg, red accent, bold type.'),

    # Day 3 — Pattern observation + morning screen
    (3, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"A pattern I\'ve noticed: traders who survive bear markets don\'t have the best entries. They have the best risk management."',
     'Authority', 'Tweet', 'Save bait', 'N', '', 'DRAFT', 'Matt Gray "pattern I noticed" formula.'),
    (3, 'Afternoon', 'X/Twitter', 'All', 'Value',
     '"My screen at 6am vs my screen at 6pm. The difference is the system." — Show workflow contrast.',
     'Reach', 'Tweet + image', 'Engagement', 'N', '', 'DRAFT', 'Visual content. AI workflow formula that got 14K views.'),
    (3, 'Morning', 'TikTok', 'All', 'Personality',
     'Before/after: cluttered charts vs clean dashboard. "This was me 6 months ago vs now." 45s screen rec.',
     'Reach', 'TikTok video', 'Follow for Part 2', 'N', '', 'DRAFT', 'Screen recording. Simplicity transition.'),

    # Day 4 — Relatable wisdom + engagement
    (4, 'Morning', 'X/Twitter', 'All', 'Personality',
     '"Everyone wants passive income from crypto. Nobody wants to stop gambling long enough to build it."',
     'Reach', 'Tweet', 'Reply bait', 'N', '', 'DRAFT', 'Farmercist formula. Motivational + crypto culture blend.'),
    (4, 'Evening', 'X/Twitter', 'All', 'Engagement',
     '"If you could only hold 3 tokens for the next 12 months — no selling, no trading — what are they and why?"',
     'Community', 'Tweet (poll/question)', 'Replies', 'N', '', 'DRAFT', 'Portfolio questions = massive replies.'),

    # Day 5 — Morning AI routine
    (5, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"I wake up to this every morning. My AI scans 17 crypto pairs and 8 forex pairs while I sleep. Not a flex — this is what replaced my 4am chart sessions."',
     'Authority', 'Tweet + screenshot', 'Engagement', 'N', '', 'DRAFT', '"I built this" + morning routine = 14K view formula.'),
    (5, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"If you stopped checking charts for 30 days, would your portfolio be fine or would you start panic-selling by day 3? Be honest."',
     'Community', 'Tweet', 'Replies', 'N', '', 'DRAFT', 'Diagnostic questions. High reply rate.'),
    (5, 'Morning', 'LinkedIn', 'All', 'Authority',
     '"Every morning, I wake up to a dashboard that already did 4 hours of work. Here\'s why I stopped grinding and started engineering."',
     'Authority', 'Long-form post', 'DM AUTOMATE', 'N', '', 'DRAFT', 'Founder framing for professional audience.'),
    (5, 'Morning', 'Instagram', 'All', 'Personality',
     'Morning routine reel: phone shows signal dashboard at 6:17am → coffee → plan day. Dark moody aesthetic. 45s.',
     'Reach', 'Reel', 'DM MORNING', 'N', '', 'DRAFT', ''),

    # Day 6 — Lesson learned + community touch
    (6, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"A lesson I wish I learned earlier: more indicators doesn\'t mean more clarity. It means more noise."',
     'Authority', 'Tweet', 'Save bait', 'N', '', 'DRAFT', '"Lesson I wish I learned earlier" = high save rate.'),
    (6, 'Evening', 'X/Twitter', 'All', 'Engagement',
     '"Drop your biggest W from Q1 2026. I\'ll start: I automated my entire content pipeline and it runs 24/7 without me touching it."',
     'Community', 'Tweet', 'Replies', 'N', '', 'DRAFT', '"Drop your W" = massive replies. Organic ContentBrain proof.'),
    (6, 'Morning', 'TikTok', 'All', 'Value',
     '8 indicators → chaos. 3 confluences → clarity. "This is why you\'re losing money." 40s screen rec.',
     'Reach', 'TikTok video', 'Comment SIMPLE', 'N', '', 'DRAFT', 'Clean chart vs cluttered chart visual.'),

    # Day 7 — Identity shift
    (7, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"3 types of people in crypto: Degens ask \'what\'s the next 100x?\' Traders ask \'what\'s the setup?\' Architects ask \'what runs while I sleep?\'"',
     'Authority', 'Tweet', 'Save/share', 'N', '', 'DRAFT', 'Identity shift content. "Architect" framing = shares + saves.'),
    (7, 'Afternoon', 'X/Twitter', 'All', 'Personality',
     '"The hardest part of building something real isn\'t the work. It\'s staying quiet while everyone else is chasing trends."',
     'Reach', 'Tweet', 'Engagement', 'N', '', 'DRAFT', 'Builder identity. Relatable.'),
    (7, 'Morning', 'LinkedIn', 'All', 'Authority',
     '"There are 3 types of founders in Web3. Only one of them builds something that lasts." — Degen/Trader/Architect framework.',
     'Authority', 'Long-form post', 'Comment ARCHITECT', 'N', '', 'DRAFT', 'Professional reframe.'),
    (7, 'Morning', 'Instagram', 'All', 'Authority',
     '"3 types of people in crypto" — Carousel (5 slides): Degen / Trader / Architect. Dark bg, red accent.',
     'Reach', 'Carousel', 'Save + tag', 'N', '', 'DRAFT', ''),

    # ── WEEK 2: AUTHORITY + COURSE TEASER (Days 8–14) ─────────────────────────

    # Day 8 — Data-backed authority + EP1 weekly series tease
    (8, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"I analyzed my last 200 trades. Here\'s the pattern that separated the winners from the losers." — Data thread.',
     'Authority', 'Thread (5-6 tweets)', 'Saves', 'N', 'Weekly Series EP1 tease', 'DRAFT',
     'First thread in weeks. Algorithm will test aggressively.'),
    (8, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"What\'s the most expensive lesson crypto taught you? Mine was trusting influencers over charts."',
     'Community', 'Tweet', 'Replies', 'N', '', 'DRAFT', 'Loss/lesson questions outperform gain questions.'),
    (8, 'Morning', 'TikTok', 'All', 'Authority',
     '"I tracked 200 trades. Here\'s the truth." — 50s spreadsheet walkthrough. R:R insight reveal.',
     'Reach', 'TikTok video', 'Comment DATA', 'N', '', 'DRAFT', 'Data walkthrough. High value.'),

    # Day 9 — WEEKLY SERIES EP1: "Things You Didn't Know You Could Do With AI"
    (9, 'Morning', 'X/Twitter', 'All', 'Weekly Series',
     'SERIES EP1: "You can ask Claude to scan your last 30 trades and tell you your worst habit. Here\'s how." — Step-by-step with screenshot.',
     'Authority', 'Thread (4 tweets)', 'Save + reply', 'N', 'Weekly Series EP1', 'DRAFT',
     'Tool: Claude. Use case: trade journal analysis. Step demo. Lead-in: "This is module 1 of something I\'m building."'),
    (9, 'Morning', 'Instagram', 'All', 'Weekly Series',
     'SERIES EP1: "Things you didn\'t know Claude could do — #1." Reel showing trade journal upload → Claude analysis → pattern output.',
     'Authority', 'Reel (45s)', 'DM for more', 'N', 'Weekly Series EP1', 'DRAFT', ''),
    (9, 'Morning', 'LinkedIn', 'All', 'Value',
     '"I spent 2 hours building a trade analytics bot with Claude. Here\'s what it found in 200 trades that I missed manually."',
     'Authority', 'Long-form post', 'DM ANALYSIS', 'Y', '', 'DRAFT', 'Course lead-in. Subtle monetization.'),

    # Day 10 — AI video workflow reveal (Creator segment)
    (10, 'Morning', 'X/Twitter', 'Creators', 'Value',
     '"I made a 45-second video this week. Total time: 23 minutes. Here\'s the exact workflow." — Nano Banana → Kling → MiniMax → CapCut.',
     'Authority', 'Thread (4 tweets)', 'Save', 'N', '', 'DRAFT', 'Creator track preview. Show the ContentBrain stack.'),
    (10, 'Morning', 'TikTok', 'Creators', 'Value',
     '"23 minutes to make a full AI video. Step by step." — Screen recording of full workflow.',
     'Reach', 'TikTok video', 'Follow for more', 'N', '', 'DRAFT', 'Tutorial format. Speed + clarity.'),
    (10, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"Which of these would you want to learn first? A) AI video creation B) AI automation C) Building tools with AI" — informal poll.',
     'Community', 'Tweet', 'Replies (validation)', 'N', '', 'DRAFT', 'Organic validation for course topic interest.'),

    # Day 11 — Business owner automation ROI story
    (11, 'Morning', 'X/Twitter', 'Business Owners', 'Authority',
     '"I automated 3 things in my business last month. The total time I got back: 14 hours/week. Here\'s exactly what they were."',
     'Authority', 'Thread (5 tweets)', 'Saves', 'Y', '', 'DRAFT', 'Business owner segment. Automation ROI framing.'),
    (11, 'Morning', 'LinkedIn', 'Business Owners', 'Authority',
     '"14 hours/week. That\'s what I got back after automating 3 business processes with AI. No developer. No agency. Just Claude and Make."',
     'Authority', 'Long-form post', 'DM AUTOMATE', 'Y', '', 'DRAFT', 'Strong lead-in for business owner audience.'),
    (11, 'Morning', 'Instagram', 'Business Owners', 'Value',
     '"What 14 hours/week back looks like" — Carousel (6 slides): the 3 automations, time saved each, tools used.',
     'Authority', 'Carousel', 'DM AUTOMATE', 'Y', '', 'DRAFT', ''),

    # Day 12 — Beginners angle: "You don't need to code"
    (12, 'Morning', 'X/Twitter', 'Beginners', 'Value',
     '"You don\'t need to know how to code to use AI. I spent 18 months thinking you did. Here\'s what I learned."',
     'Reach', 'Thread (3 tweets)', 'Save', 'N', '', 'DRAFT', 'Beginners segment. Lower barrier messaging.'),
    (12, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"Be honest: what\'s the #1 thing stopping you from using AI in your work or business right now?"',
     'Community', 'Tweet', 'Replies (pain points)', 'N', '', 'DRAFT', 'Market research post. Read replies as demand signal.'),
    (12, 'Morning', 'TikTok', 'Beginners', 'Value',
     '"I thought you needed to code to use AI. I was wrong." — 35s talking head. 3 things you can do with AI right now, zero coding.',
     'Reach', 'TikTok video', 'Follow for more', 'N', '', 'DRAFT', 'Beginners reassurance. High mass appeal.'),

    # Day 13 — FREEBIE DROP + weekend engagement
    (13, 'Morning', 'X/Twitter', 'All', 'Freebie',
     '"Free resource: The AI Tool Stack I Use Every Day. 8 tools. What each one does. When to use which." — Drop as image/PDF thread.',
     'Leads', 'Thread + image', 'RT + save', 'Y', '', 'DRAFT', 'Freebie = algorithm push + email list growth potential.'),
    (13, 'Morning', 'Telegram', 'All', 'Freebie',
     '"AI Tool Stack guide — FREE. The 8 tools I run my entire business on. Download below."',
     'Community', 'Community post', 'Download CTA', 'Y', '', 'DRAFT', 'Telegram community activation.'),
    (13, 'Morning', 'Instagram', 'All', 'Freebie',
     '"My complete AI tool stack — free breakdown." Carousel (8 slides): one tool per slide, what it does, when to use.',
     'Reach', 'Carousel', 'Save + share', 'N', '', 'DRAFT', ''),

    # Day 14 — Transformational
    (14, 'Morning', 'X/Twitter', 'All', 'Transformational',
     '"What changed when I stopped doing things manually: my mornings got 2 hours back. My content improved. My anxiety about consistency disappeared."',
     'Reach', 'Thread (4 tweets)', 'Save', 'N', '', 'DRAFT', 'Transformation arc. Before/after without pitching.'),
    (14, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"Hot take: most people don\'t need more tools. They need to use 3 tools properly. Agree or disagree?"',
     'Community', 'Tweet', 'Replies', 'N', '', 'DRAFT', 'Weekend hot take.'),

    # ── WEEK 3: COURSE ANNOUNCEMENT + VALIDATION (Days 15–21) ─────────────────

    # Day 15 — VALIDATION POST (7 days from today)
    (15, 'Morning', 'X/Twitter', 'All', 'Lead Gen',
     'VALIDATION: "Quick question: if I ran live AI classes twice a week on Telegram + Google Meet — $30 for the full month — would you join? What topic first: A) AI Automation B) AI Video C) Coding with AI"',
     'Leads', 'Tweet (poll)', 'Reply validation', 'Y', 'Validation Post', 'DRAFT',
     'Publish exactly Day 15 (April 17). Measure real interest before committing to full launch.'),
    (15, 'Morning', 'LinkedIn', 'All', 'Lead Gen',
     'VALIDATION: "I\'m considering running live AI classes for founders and creators — $30/month, live twice a week. Would this be useful to you? What would make it worth it?"',
     'Leads', 'Post + poll', 'Comments', 'Y', 'Validation Post', 'DRAFT', 'LinkedIn version — professional angle.'),
    (15, 'Morning', 'Instagram', 'All', 'Lead Gen',
     'VALIDATION: "Would you pay $30 for live AI classes twice a week?" — Story poll + reel version.',
     'Leads', 'Story poll + Reel', 'Swipe + vote', 'Y', 'Validation Post', 'DRAFT', 'Instagram version.'),
    (15, 'Morning', 'TikTok', 'All', 'Lead Gen',
     'VALIDATION: "I\'m thinking about running a live AI course. Would you join? React if yes." — Talking head 30s.',
     'Leads', 'TikTok video', 'React/comment', 'Y', 'Validation Post', 'DRAFT', ''),

    # Day 16 — Money proof content
    (16, 'Morning', 'X/Twitter', 'All', 'Money Proof',
     'INCOME BREAKDOWN: "5 ways people are making real money with AI tools in 2026. With numbers." — Thread.',
     'Authority', 'Thread (6 tweets)', 'Save + share', 'Y', 'Money Proof EP1', 'DRAFT',
     'Income streams: freelancing $2-8K/mo, Claude Code projects $500-10K/mo, automation services $800+$200/mo, AI video per deal $200-5K, AI consulting $5K-50K/mo. Source every claim.'),
    (16, 'Morning', 'LinkedIn', 'Business Owners', 'Money Proof',
     '"What AI consultants are charging in 2026. And how beginners are already starting at $500/month." — Income data post.',
     'Authority', 'Long-form post', 'DM for more', 'Y', 'Money Proof EP1', 'DRAFT', ''),
    (16, 'Morning', 'Instagram', 'All', 'Money Proof',
     '"How people are making money with AI — 5 real income streams." Carousel (6 slides). Numbers sourced.',
     'Authority', 'Carousel', 'Save + share', 'Y', 'Money Proof EP1', 'DRAFT', ''),

    # Day 17 — Course announcement
    (17, 'Morning', 'X/Twitter', 'All', 'Course Launch',
     '"Something I\'ve been building for the past 30 days. A live AI course. 3 topics. 4 audience tracks. $15–$50. Launching May 3. Here\'s what\'s inside." — Announcement thread.',
     'Leads', 'Thread (5 tweets)', 'DM COURSE', 'Y', 'Course Announcement', 'DRAFT',
     'First hard announcement. Clear, specific, no fluff.'),
    (17, 'Morning', 'LinkedIn', 'All', 'Course Launch',
     '"I\'m launching an AI course in 16 days. Live classes. 4 audience tracks. Priced for accessibility. Here\'s why I built it." — Announcement post.',
     'Leads', 'Long-form post', 'DM COURSE', 'Y', 'Course Announcement', 'DRAFT', ''),
    (17, 'Morning', 'Telegram', 'All', 'Course Launch',
     '"IT\'S LIVE. Quivira AI Course announced. 3 topics, 4 tracks, 30 days. Link inside." — Community announcement.',
     'Community', 'Community post', 'Click link', 'Y', 'Course Announcement', 'DRAFT', ''),
    (17, 'Morning', 'Instagram', 'All', 'Course Launch',
     '"The course is coming." — Reel reveal. Dark, cinematic. Show the framework, the tools, the stack. No price yet. Mysterious.',
     'Reach', 'Reel (30s)', 'DM COURSE', 'Y', 'Course Announcement', 'DRAFT', ''),

    # Day 18 — AI Automation deep dive (value to drive course interest)
    (18, 'Morning', 'X/Twitter', 'Beginners', 'Value',
     '"How I set up my first automation in 47 minutes with zero prior experience. The exact steps." — Tutorial thread.',
     'Authority', 'Thread (5 tweets)', 'Save + DM', 'Y', '', 'DRAFT', 'Beginners use case. Course content preview.'),
    (18, 'Morning', 'TikTok', 'Beginners', 'Value',
     '"Zero experience → first automation in 47 minutes. Watch me do it." — Speed run screen recording.',
     'Reach', 'TikTok video', 'Comment AUTOMATE', 'Y', '', 'DRAFT', ''),
    (18, 'Afternoon', 'X/Twitter', 'All', 'Lead Gen',
     '"Early access to the Quivira AI Course is now open. 30 days live. $15–$50 depending on tier. DM me COURSE for the link."',
     'Leads', 'Tweet', 'DM COURSE', 'Y', '', 'DRAFT', 'Course CTA. Short and direct.'),

    # Day 19 — WEEKLY SERIES EP2: AI Video workflow (Creators)
    (19, 'Morning', 'X/Twitter', 'Creators', 'Weekly Series',
     'SERIES EP2: "You can generate a cinematic 5-second video from a single text prompt using Kling. Here\'s how." — Step + demo.',
     'Authority', 'Thread (4 tweets)', 'Save + DM', 'Y', 'Weekly Series EP2', 'DRAFT',
     'Tool: Kling AI. Use case: social media video creation. Lead-in to AI Video Creation course.'),
    (19, 'Morning', 'Instagram', 'Creators', 'Weekly Series',
     'SERIES EP2: "Things you didn\'t know Kling could do — #2." Reel showing image → 5s video → platform-ready clip.',
     'Reach', 'Reel (45s)', 'DM KLING', 'Y', 'Weekly Series EP2', 'DRAFT', ''),
    (19, 'Morning', 'LinkedIn', 'Creators', 'Value',
     '"Why every content creator should have an AI video workflow in 2026. And why most don\'t yet." — Industry insight.',
     'Authority', 'Long-form post', 'DM VIDEO', 'Y', '', 'DRAFT', ''),

    # Day 20 — Developer angle: "Built with AI in 4 hours"
    (20, 'Morning', 'X/Twitter', 'Developers', 'Authority',
     '"I built a working web app in 4 hours using only Claude Code and zero prior framework knowledge. Here\'s what happened." — Build thread.',
     'Authority', 'Thread (5 tweets)', 'Save', 'Y', '', 'DRAFT', 'Developers segment. Vibe coding credibility.'),
    (20, 'Morning', 'TikTok', 'Developers', 'Value',
     '"I built a crypto dashboard in 4 hours with Claude Code. Zero frameworks. Watch the timelapse." — Screen recording.',
     'Reach', 'TikTok video', 'Comment BUILD', 'Y', '', 'DRAFT', ''),

    # Day 21 — Lead gen + course CTA
    (21, 'Morning', 'X/Twitter', 'All', 'Lead Gen',
     '"The Quivira AI Course has 3 tiers: $15 (self-paced), $30 (live classes), $50 (full access + direct DM to me + 2 bonuses). Which one is you? DM me the price."',
     'Leads', 'Tweet', 'DM tier', 'Y', '', 'DRAFT', 'Tiered pricing reveal. Simple, clear, no fluff.'),
    (21, 'Morning', 'LinkedIn', 'All', 'Lead Gen',
     '"I\'m running this course at founding prices because I want proof this works. Here\'s what each tier gets you." — Pricing breakdown post.',
     'Leads', 'Long-form post', 'DM PRICING', 'Y', '', 'DRAFT', ''),
    (21, 'Morning', 'Telegram', 'All', 'Lead Gen',
     '"Seats are limited. Full Access (all 3 topics + direct DM access) is capped at 20. Here\'s the link."',
     'Community', 'Community post', 'Click link', 'Y', '', 'DRAFT', ''),

    # ── WEEK 4: ACTIVE PROMOTION + SOCIAL PROOF (Days 22–28) ─────────────────

    # Day 22 — Money proof post EP2
    (22, 'Morning', 'X/Twitter', 'All', 'Money Proof',
     'CASE STUDY: "This person with zero coding experience built an automation that saves their agency $2,000/month. Here\'s how they did it." — Breakdown thread.',
     'Authority', 'Thread (4 tweets)', 'Save + share', 'Y', 'Money Proof EP2', 'DRAFT',
     'Use industry benchmark examples clearly labeled. No invented data.'),
    (22, 'Morning', 'LinkedIn', 'Business Owners', 'Money Proof',
     '"How a non-technical founder saved $2,000/month using AI automation. And what it cost to set up: $0 beyond tool subscriptions."',
     'Authority', 'Long-form post', 'DM for more', 'Y', 'Money Proof EP2', 'DRAFT', ''),
    (22, 'Afternoon', 'X/Twitter', 'All', 'Course Launch',
     '"8 days left. Quivira AI Course. $15–$50. Live classes start May 3. DM me COURSE."',
     'Leads', 'Tweet', 'DM COURSE', 'Y', '', 'DRAFT', 'Countdown urgency.'),

    # Day 23 — Social proof / early student experience
    (23, 'Morning', 'X/Twitter', 'All', 'Course Launch',
     '"First 5 students are in. Here\'s what they said when they saw the full curriculum." — Screenshot-style quote thread.',
     'Leads', 'Thread (3 tweets)', 'DM COURSE', 'Y', 'Social Proof', 'DRAFT',
     'If actual student quotes not available, use validation poll results as social proof.'),
    (23, 'Morning', 'Instagram', 'All', 'Course Launch',
     '"What\'s inside the Quivira AI Course" — Carousel walkthrough (8 slides): topics, tracks, live sessions, bonuses.',
     'Leads', 'Carousel', 'DM COURSE', 'Y', '', 'DRAFT', ''),
    (23, 'Morning', 'TikTok', 'All', 'Course Launch',
     '"I\'m running a live AI course for 30 days. Here\'s what Week 1 looks like inside." — Behind the scenes footage.',
     'Leads', 'TikTok video', 'Link in bio', 'Y', '', 'DRAFT', ''),

    # Day 24 — FREEBIE + WEEKLY SERIES EP3
    (24, 'Morning', 'X/Twitter', 'All', 'Freebie',
     'SERIES EP3: "You can use Supabase + Claude to build a database that queries itself in plain English. No SQL needed. Here\'s how." + FREE: Module 1 pre-recorded (24hr access).',
     'Leads', 'Thread (4 tweets)', 'DM MODULE1', 'Y', 'Weekly Series EP3', 'DRAFT',
     'Tool: Supabase + Claude. Double value: series + freebie. Drives course interest.'),
    (24, 'Morning', 'Telegram', 'All', 'Freebie',
     '"FREE for 24 hours: Module 1 of the AI Automation course. Pre-recorded. Zero payment required." — Link drop.',
     'Leads', 'Community post', 'Click link', 'Y', 'Weekly Series EP3', 'DRAFT', ''),
    (24, 'Morning', 'Instagram', 'All', 'Weekly Series',
     'SERIES EP3: "Things you didn\'t know Supabase + Claude could do." Reel showing natural language → database query output.',
     'Authority', 'Reel (45s)', 'DM MODULE1', 'Y', 'Weekly Series EP3', 'DRAFT', ''),

    # Day 25 — Course behind the scenes
    (25, 'Morning', 'X/Twitter', 'All', 'Personality',
     '"Here\'s what Week 1 inside the Quivira AI Course looks like. Real screenshots. Real builds. Real feedback." — Course BTS thread.',
     'Authority', 'Thread (4 tweets)', 'DM COURSE', 'Y', '', 'DRAFT', ''),
    (25, 'Morning', 'TikTok', 'All', 'Course Launch',
     '"Week 1 inside my AI course — what we built in 7 days." — Screen recording montage of live sessions.',
     'Leads', 'TikTok video', 'Link in bio', 'Y', '', 'DRAFT', ''),

    # Day 26 — Authority: AI skills for 2027
    (26, 'Morning', 'X/Twitter', 'All', 'Authority',
     '"The AI skills that will matter most in 2027: 1) Building with AI (not just using it) 2) Workflow architecture 3) Prompt engineering 4) AI video production 5) Multi-agent systems." — Thread.',
     'Authority', 'Thread (5 tweets)', 'Save', 'N', '', 'DRAFT', 'Authority piece. Not promotional.'),
    (26, 'Morning', 'LinkedIn', 'All', 'Authority',
     '"The 5 AI skills companies will pay premium for in 2027. And how to build all of them in 30 days." — Industry authority post.',
     'Authority', 'Long-form post', 'DM for more', 'Y', '', 'DRAFT', ''),

    # Day 27 — Urgency: 3 days left
    (27, 'Morning', 'X/Twitter', 'All', 'Course Launch',
     '"3 days left to join Quivira AI Course at founding prices. After May 3, prices go up. $15 / $30 / $50 — pick your tier."',
     'Leads', 'Tweet', 'DM COURSE', 'Y', '', 'DRAFT', 'Genuine urgency — cohort 1 pricing is locked.'),
    (27, 'Afternoon', 'X/Twitter', 'All', 'Engagement',
     '"What would stop you from joining a live AI course right now? I\'ll answer every reply honestly."',
     'Community', 'Tweet', 'Replies (objections)', 'Y', '', 'DRAFT', 'Objection handling thread. Read replies carefully.'),
    (27, 'Morning', 'LinkedIn', 'All', 'Course Launch',
     '"3 days left. If you\'ve been watching from the sidelines — this is the moment." — Final pitch LinkedIn.',
     'Leads', 'Post', 'DM COURSE', 'Y', '', 'DRAFT', ''),
    (27, 'Morning', 'Telegram', 'All', 'Course Launch',
     '"3 days. 3 days until founding prices close. Who\'s still on the fence? Drop your question here."',
     'Community', 'Community post', 'Replies', 'Y', '', 'DRAFT', ''),

    # Day 28 — Community wins
    (28, 'Morning', 'X/Twitter', 'All', 'Transformational',
     '"Week 1 win from inside the course: one student automated their entire client onboarding process in under 90 minutes. Zero prior AI experience." — Win showcase.',
     'Leads', 'Tweet', 'DM COURSE', 'Y', 'Social Proof', 'DRAFT', 'Social proof. Real wins from inside the cohort.'),
    (28, 'Morning', 'Instagram', 'All', 'Transformational',
     '"What students are building inside the Quivira AI Course — Week 1 results." Reel montage or carousel of wins.',
     'Leads', 'Reel', 'DM COURSE', 'Y', 'Social Proof', 'DRAFT', ''),

    # ── DAYS 29–30: FINAL PUSH ────────────────────────────────────────────────

    # Day 29 — FOMO
    (29, 'Morning', 'X/Twitter', 'All', 'Course Launch',
     '"Last 24 hours. Quivira AI Course Cohort 1. After tomorrow, we close enrollments and prices go up for Cohort 2." — Final day countdown.',
     'Leads', 'Thread (3 tweets)', 'DM COURSE', 'Y', '', 'DRAFT', 'Genuine scarcity. Cohort 1 is limited.'),
    (29, 'Afternoon', 'X/Twitter', 'All', 'Lead Gen',
     '"Full Access tier (all 3 topics + direct DM access) — X seats remaining. Once gone, gone." — Seat counter.',
     'Leads', 'Tweet', 'DM FULLACCESS', 'Y', '', 'DRAFT', 'Real seat scarcity drives action.'),
    (29, 'Morning', 'LinkedIn', 'All', 'Course Launch',
     '"Today is the last day to join Cohort 1 at founding prices. What you\'re getting for $50 is worth 10× that." — Final LinkedIn push.',
     'Leads', 'Post', 'DM COURSE', 'Y', '', 'DRAFT', ''),
    (29, 'Morning', 'Telegram', 'All', 'Course Launch',
     '"LAST DAY. Cohort 1 closes tonight. Link inside." — Pin this message.',
     'Community', 'Community post', 'Click link', 'Y', '', 'DRAFT', 'Pin this in the Telegram group.'),
    (29, 'Morning', 'Instagram', 'All', 'Course Launch',
     '"Last 24 hours." — Countdown reel. Simple text animation on dark background. Quivira aesthetic.',
     'Leads', 'Reel (15s)', 'Link in bio', 'Y', '', 'DRAFT', ''),

    # Day 30 — FINAL DAY + WEEKLY SERIES EP4
    (30, 'Morning', 'X/Twitter', 'All', 'Course Launch',
     '"Today is the last day. Here\'s everything you get inside the Quivira AI Course." — Full summary thread.',
     'Leads', 'Thread (6 tweets)', 'DM COURSE', 'Y', '', 'DRAFT',
     'Final pitch. Clear, complete, confident. No desperation.'),
    (30, 'Afternoon', 'X/Twitter', 'All', 'Weekly Series',
     'SERIES EP4: "You can use Remotion to programmatically generate 30 videos from a CSV file. Here\'s the exact script." — Final series ep.',
     'Authority', 'Thread (4 tweets)', 'DM REMOTION', 'Y', 'Weekly Series EP4', 'DRAFT',
     'Tool: Remotion. Use case: batch video generation. Lead-in: "We teach this in Week 3 of Coding with AI."'),
    (30, 'Morning', 'LinkedIn', 'All', 'Personality',
     '"Today Cohort 1 closes. Thank you to everyone who joined. Here\'s what we\'re going to build together over the next 30 days." — Gratitude + hype.',
     'Community', 'Post', 'DM COURSE', 'Y', '', 'DRAFT', ''),
    (30, 'Morning', 'Instagram', 'All', 'Weekly Series',
     'SERIES EP4: "Things you didn\'t know Remotion could do." Reel: CSV → 30 videos generated automatically.',
     'Authority', 'Reel (45s)', 'DM REMOTION', 'Y', 'Weekly Series EP4', 'DRAFT', ''),
    (30, 'Morning', 'TikTok', 'All', 'Course Launch',
     '"The Quivira AI Course starts tomorrow. Here\'s what Cohort 1 looks like." — Final hype video.',
     'Leads', 'TikTok video', 'Link in bio', 'Y', '', 'DRAFT', ''),
]

COLUMNS = [
    'Day #', 'Date', 'Day of Week', 'Week', 'Post Slot', 'Platform',
    'Audience Segment', 'Content Category', 'Hook / Topic',
    'Content Goal', 'Format', 'CTA', 'Monetization',
    'Series Tag', 'Status', 'Notes'
]

COL_WIDTHS = [6, 11, 12, 7, 10, 13, 16, 16, 60, 13, 18, 20, 13, 18, 11, 40]


def week_of(day_num):
    if day_num <= 7: return 1
    if day_num <= 14: return 2
    if day_num <= 21: return 3
    if day_num <= 28: return 4
    return 5


def build_calendar():
    wb = Workbook()
    ws = wb.active
    ws.title = '30-Day Launch Calendar'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Title row
    ws.merge_cells('A1:P1')
    title = ws['A1']
    title.value = 'QUIVIRA AI COURSE — 30-Day Launch Calendar  |  April 3 – May 2, 2026  |  Day 1 = LOCKED (already posted)'
    title.fill = fill(C['dark'])
    title.font = font(C['white'], True, 11)
    title.alignment = align('center', 'center', False)
    ws.row_dimensions[1].height = 22

    # Header row
    ws.row_dimensions[2].height = 18
    for col_idx, (col_name, col_w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=col_idx)
        style_header(cell, col_name, C['dark'], C['white'])
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    # Data rows
    row = 3
    for post in POSTS:
        (day_num, slot, platform, segment, category,
         hook, goal, fmt, cta, monetize, series, status, notes) = post

        day_date = START + timedelta(days=day_num - 1)
        week = week_of(day_num)
        week_bg = WEEK_BG.get(week, C['white'])
        status_bg = STATUS_COLOR.get(status, C['white'])
        cat_bg = CATEGORY_COLOR.get(category, C['white'])
        plat_col = PLATFORM_COLOR.get(platform, C['dark'])

        values = [
            day_num,
            day_date.strftime('%b %d'),
            day_date.strftime('%A'),
            f'Week {week}',
            slot,
            platform,
            segment,
            category,
            hook,
            goal,
            fmt,
            cta,
            monetize,
            series,
            status,
            notes,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx)
            col_name = COLUMNS[col_idx - 1]

            if col_name == 'Platform':
                style_cell(cell, val, plat_col, C['white'], bold=True, h='center')
            elif col_name == 'Content Category':
                style_cell(cell, val, cat_bg, C['gray'], h='center')
            elif col_name == 'Status':
                style_cell(cell, val, status_bg, C['dark'], bold=(status == 'LOCKED'), h='center')
            elif col_name == 'Monetization':
                mon_bg = 'E5FFE5' if val == 'Y' else C['white']
                style_cell(cell, val, mon_bg, C['gray'], bold=(val == 'Y'), h='center')
            elif col_name in ('Hook / Topic', 'Notes'):
                style_cell(cell, val, C['white'], C['gray'], h='left', v='top', wrap=True)
            elif col_name == 'Series Tag':
                s_bg = 'CCF5FF' if val else C['white']
                style_cell(cell, val, s_bg, C['dark'], h='center')
            elif col_name == 'Day #':
                day_bg = C['locked'] if status == 'LOCKED' else week_bg
                style_cell(cell, val, day_bg, C['dark'], bold=True, h='center')
            elif col_name in ('Date', 'Day of Week', 'Week'):
                style_cell(cell, val, week_bg, C['gray'], h='center')
            else:
                style_cell(cell, val, C['white'], C['gray'], h='left')

        ws.row_dimensions[row].height = 40
        row += 1

    # Legend
    ws.append([])
    ws.append([])
    legend_row = row + 2
    ws.cell(row=legend_row, column=1).value = 'LEGEND'
    ws.cell(row=legend_row, column=1).font = font(C['dark'], True, 10)

    legend_items = [
        ('LOCKED', C['locked'], 'Day 1 — already posted, do not modify'),
        ('POSTED', C['posted'], 'Content confirmed published'),
        ('SCHEDULED', 'E0F0FF', 'Queued in Typefully / Buffer'),
        ('DRAFT', C['draft'], 'Written, not yet scheduled'),
    ]
    for i, (status_text, bg, desc) in enumerate(legend_items):
        r = legend_row + 1 + i
        ws.cell(row=r, column=1).value = status_text
        ws.cell(row=r, column=1).fill = fill(bg)
        ws.cell(row=r, column=1).font = font(C['dark'], True)
        ws.cell(row=r, column=1).border = border_thin()
        ws.cell(row=r, column=2).value = desc
        ws.cell(row=r, column=2).font = font(C['gray'])

    wb.save(CAL_PATH)
    print(f'Calendar saved: {CAL_PATH}')


# ── TO-DO TRACKER ─────────────────────────────────────────────────────────────

TODO_ITEMS = [
    # (task, category, owner, due_date_day, priority, risk_flag, fix_suggestion)
    # COURSE CREATION
    ('Write lesson scripts for AI Automation — Beginner Track (M1–M5)', 'Course Creation', '@big_quiv', 10, 'HIGH', False, ''),
    ('Write lesson scripts for AI Automation — Developer Track (M1–M6)', 'Course Creation', '@big_quiv', 12, 'HIGH', False, ''),
    ('Write lesson scripts for AI Automation — Creator Track (M1–M5)', 'Course Creation', '@big_quiv', 12, 'HIGH', False, ''),
    ('Write lesson scripts for AI Automation — Business Track (M1–M5)', 'Course Creation', '@big_quiv', 14, 'HIGH', False, ''),
    ('Write lesson scripts for AI Video Creation — all 4 tracks', 'Course Creation', '@big_quiv', 16, 'HIGH', True, 'RISK: 12 tracks of scripts in 16 days is aggressive. Prioritise Beginner + Creator tracks first — they have the largest audience. Defer Developer + Business scripts to Day 18 max.'),
    ('Write lesson scripts for Coding with AI — all 4 tracks', 'Course Creation', '@big_quiv', 18, 'HIGH', True, 'RISK: Same risk as above. Batch-write with Claude using module outlines as prompts. Target: 2 tracks/day working session.'),
    ('Create slide decks / visual aids for live sessions (Google Slides)', 'Course Creation', '@big_quiv', 20, 'MEDIUM', False, ''),
    ('Record all pre-recorded modules (M2, M5 across topics)', 'Course Creation', '@big_quiv', 24, 'HIGH', True, 'RISK: Pre-recorded modules need to be done by Day 24 minimum to release 48hrs before live sessions. Use screen recording + MiniMax voiceover for efficiency.'),
    ('Create resource PDFs: prompt templates, workflow checklists, tool guides', 'Course Creation', '@big_quiv', 22, 'MEDIUM', False, ''),
    ('Create completion certificate template (branded PDF)', 'Course Creation', '@big_quiv', 26, 'LOW', False, ''),
    ('Build Notion portal page (course hub — links to all pre-recorded modules)', 'Course Creation', '@big_quiv', 15, 'HIGH', False, ''),

    # TECH SETUP
    ('Set up Gumroad products: 3 tiers ($15 / $30 / $50)', 'Tech Setup', '@big_quiv', 13, 'HIGH', False, ''),
    ('Set up Selar.co products (Nigerian Naira buyers)', 'Tech Setup', '@big_quiv', 14, 'MEDIUM', False, ''),
    ('Create Telegram group for course community (separate from signal group)', 'Tech Setup', '@big_quiv', 8, 'HIGH', False, ''),
    ('Set up Google Meet recurring links for live sessions (Tue + Thu evenings)', 'Tech Setup', '@big_quiv', 10, 'HIGH', False, ''),
    ('Configure Notion portal with password protection and module links', 'Tech Setup', '@big_quiv', 16, 'HIGH', False, ''),
    ('Set up welcome message and onboarding flow in Telegram group', 'Tech Setup', '@big_quiv', 17, 'MEDIUM', False, ''),
    ('Test full student journey: payment → Notion access → Telegram invite → first module', 'Tech Setup', '@big_quiv', 20, 'HIGH', True, 'RISK: If this breaks on launch day, students lose confidence immediately. Run a complete dry-run by Day 20 minimum.'),
    ('Create landing page or Linktree with course links (Gumroad + Selar)', 'Tech Setup', '@big_quiv', 16, 'MEDIUM', False, ''),
    ('Set up auto-DM or ManyChat flow for keyword triggers (COURSE, STARTER, BUILDER, FULLACCESS)', 'Tech Setup', '@big_quiv', 18, 'MEDIUM', False, ''),

    # CONTENT CREATION
    ('Record Day 1 validation poll content (all 4 platforms)', 'Content Creation', '@big_quiv', 15, 'HIGH', False, ''),
    ('Write and schedule Week 1 content (Days 2–7) on Typefully', 'Content Creation', '@big_quiv', 5, 'HIGH', True, 'RISK: Days 2-7 need to be written and scheduled by Day 5. Already have outlines. Execute now.'),
    ('Write and schedule Week 2 content (Days 8–14)', 'Content Creation', '@big_quiv', 11, 'HIGH', False, ''),
    ('Write Weekly Series EP1 (Claude trade journal analysis)', 'Content Creation', '@big_quiv', 9, 'HIGH', False, ''),
    ('Write Weekly Series EP2 (Kling AI video workflow)', 'Content Creation', '@big_quiv', 19, 'HIGH', False, ''),
    ('Write Weekly Series EP3 (Supabase + Claude + freebie)', 'Content Creation', '@big_quiv', 24, 'HIGH', False, ''),
    ('Write Weekly Series EP4 (Remotion batch video generation)', 'Content Creation', '@big_quiv', 30, 'HIGH', False, ''),
    ('Create AI Tool Stack resource PDF (freebie for Day 13)', 'Content Creation', '@big_quiv', 12, 'MEDIUM', False, ''),
    ('Write course announcement thread and all platform versions', 'Content Creation', '@big_quiv', 17, 'HIGH', False, ''),
    ('Write money proof content — income breakdown thread + carousel', 'Content Creation', '@big_quiv', 16, 'HIGH', False, ''),
    ('Create carousel: "What $30 gets you in Quivira AI Course"', 'Content Creation', '@big_quiv', 21, 'MEDIUM', False, ''),
    ('Write Week 4 promotional content (Days 22–28)', 'Content Creation', '@big_quiv', 21, 'HIGH', False, ''),
    ('Write final push content (Days 29–30)', 'Content Creation', '@big_quiv', 28, 'HIGH', False, ''),

    # PROMOTION
    ('Post validation poll on all platforms (Day 15)', 'Promotion', '@big_quiv', 15, 'HIGH', False, ''),
    ('Monitor validation poll replies and DM everyone who responds', 'Promotion', '@big_quiv', 16, 'HIGH', False, ''),
    ('Drop free Module 1 on Telegram + X (Day 24)', 'Promotion', '@big_quiv', 24, 'MEDIUM', False, ''),
    ('Reach out to 5 relevant creators for collaboration / cross-promotion', 'Promotion', '@big_quiv', 20, 'MEDIUM', True, 'RISK: Organic only — this may not yield results in time. Start outreach by Day 12 minimum to allow response time.'),
    ('Pin course announcement in Telegram group', 'Promotion', '@big_quiv', 17, 'HIGH', False, ''),
    ('DM follow-up to everyone who engaged with validation poll', 'Promotion', '@big_quiv', 18, 'HIGH', False, ''),
    ('Create urgency posts for Days 27–29 (3-day countdown)', 'Promotion', '@big_quiv', 26, 'HIGH', False, ''),
    ('Collect student testimonials from Week 1 and publish as social proof', 'Promotion', '@big_quiv', 25, 'MEDIUM', False, ''),

    # COMMUNITY SETUP
    ('Write Telegram welcome message for course community', 'Community Setup', '@big_quiv', 8, 'HIGH', False, ''),
    ('Write community rules + FAQ document', 'Community Setup', '@big_quiv', 10, 'MEDIUM', False, ''),
    ('Create onboarding checklist for new students (pinned Notion doc)', 'Community Setup', '@big_quiv', 14, 'MEDIUM', False, ''),
    ('Set up weekly check-in cadence in Telegram group', 'Community Setup', '@big_quiv', 17, 'MEDIUM', False, ''),
    ('Create "before you start" resource document for each track', 'Community Setup', '@big_quiv', 20, 'LOW', False, ''),

    # CLASS DELIVERY
    ('Prepare and rehearse M1 live session for all 3 topics', 'Class Delivery', '@big_quiv', 26, 'HIGH', True, 'RISK: First live session (Apr 28 or May 1) needs a full dry-run. Allocate 2 hrs for rehearsal on Day 26.'),
    ('Test Google Meet screen share, audio, and recording', 'Class Delivery', '@big_quiv', 18, 'HIGH', False, ''),
    ('Create session agenda template for all live classes', 'Class Delivery', '@big_quiv', 22, 'MEDIUM', False, ''),
    ('Set up recording workflow: Google Meet → Drive → Notion link', 'Class Delivery', '@big_quiv', 20, 'HIGH', False, ''),
    ('Prepare Q&A protocols: how to handle common questions live', 'Class Delivery', '@big_quiv', 24, 'LOW', False, ''),
    ('Draft completion certificate and test PDF generation', 'Class Delivery', '@big_quiv', 28, 'LOW', False, ''),
]

TODO_COLUMNS = [
    'Task', 'Category', 'Owner', 'Due Date', 'Day #', 'Priority',
    'Status', 'Risk Flag', 'Risk / Fix Suggestion'
]
TODO_COL_WIDTHS = [55, 16, 12, 13, 7, 10, 12, 10, 60]

CAT_COLORS_TODO = {
    'Course Creation':  'EAF4FF',
    'Tech Setup':       'FFF5E6',
    'Content Creation': 'EAF5EA',
    'Promotion':        'FFF0F0',
    'Community Setup':  'F5E5FF',
    'Class Delivery':   'FFF5CC',
}

PRIORITY_COLORS = {
    'HIGH':   'FFD5D5',
    'MEDIUM': 'FFFACC',
    'LOW':    'E5F5E5',
}


def build_todo():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Launch To-Do Tracker'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(TODO_COLUMNS))}1')
    title = ws['A1']
    title.value = 'QUIVIRA AI COURSE — Launch To-Do Tracker  |  30-Day Window  |  April 3 – May 2, 2026'
    title.fill = fill(C['dark'])
    title.font = font(C['white'], True, 11)
    title.alignment = align('center', 'center', False)
    ws.row_dimensions[1].height = 22

    # Headers
    ws.row_dimensions[2].height = 18
    for col_idx, (col_name, col_w) in enumerate(zip(TODO_COLUMNS, TODO_COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=col_idx)
        style_header(cell, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    # Data
    row = 3
    for item in TODO_ITEMS:
        (task, category, owner, due_day, priority, risk_flag, fix) = item
        due_date = START + timedelta(days=due_day - 1)
        cat_bg = CAT_COLORS_TODO.get(category, C['white'])
        pri_bg = PRIORITY_COLORS.get(priority, C['white'])

        values = [
            task, category, owner,
            due_date.strftime('%b %d, %Y'),
            due_day, priority, 'To Do',
            '⚠ YES' if risk_flag else 'OK',
            fix if risk_flag else ''
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx)
            col_name = TODO_COLUMNS[col_idx - 1]

            if col_name == 'Category':
                style_cell(cell, val, cat_bg, C['gray'], h='center')
            elif col_name == 'Priority':
                style_cell(cell, val, pri_bg, C['dark'], bold=True, h='center')
            elif col_name == 'Risk Flag':
                rf_bg = 'FFD5D5' if risk_flag else 'E5F5E5'
                rf_col = C['rose'] if risk_flag else C['green']
                style_cell(cell, val, rf_bg, rf_col, bold=risk_flag, h='center')
            elif col_name == 'Status':
                style_cell(cell, val, C['draft'], C['dark'], h='center')
            elif col_name in ('Task', 'Risk / Fix Suggestion'):
                style_cell(cell, val, C['white'], C['gray'], h='left', wrap=True)
            else:
                style_cell(cell, val, C['white'], C['gray'], h='center')

        ws.row_dimensions[row].height = 35
        row += 1

    wb.save(TODO_PATH)
    print(f'To-do tracker saved: {TODO_PATH}')


if __name__ == '__main__':
    build_calendar()
    build_todo()
