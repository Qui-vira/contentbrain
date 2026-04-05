"""
Rewrite the 30-day launch calendar with balanced audience segments.
Overwrites 06-Drafts/30-day-launch-calendar.xlsx in place.
Day 1 is locked. Days 2-30 are rebuilt with all 4 segments represented
every week, no segment dominating more than 2 consecutive days.
"""
import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', '06-Drafts', '30-day-launch-calendar.xlsx')
START = date(2026, 4, 3)

# ── PALETTE ───────────────────────────────────────────────────────────────────
C = {
    'dark': '1A1A2E', 'red': 'E94560', 'gray': '444444',
    'light': 'F5F5F5', 'white': 'FFFFFF', 'mid': 'CCCCCC',
    'green': '2D6A4F', 'blue': '1D4E89', 'purple': '6B3FA0',
    'teal': '1A6B6B', 'rose': 'C0392B',
    'posted': 'D5F0D5', 'draft': 'FFF5D5', 'locked': 'E8E8E8',
}

WEEK_BG = {1: 'EAF4FF', 2: 'FFF5E6', 3: 'EAF5EA', 4: 'FFF0F0', 5: 'F0E6FF'}

PLATFORM_COLOR = {
    'X/Twitter': C['dark'], 'LinkedIn': C['blue'],
    'TikTok': C['teal'], 'Instagram': C['rose'],
    'Telegram': C['green'], 'All': C['purple'],
}

CATEGORY_COLOR = {
    'Personality': 'FFE5EC', 'Value': 'E5F5FF', 'Authority': 'FFF5CC',
    'Freebie': 'E5FFE5', 'Transformational': 'F5E5FF', 'Lead Gen': 'FFE8CC',
    'Weekly Series': 'CCF5FF', 'Money Proof': 'FFFACC',
    'Course Launch': 'FFCCCC', 'Engagement': 'EEEEEE',
}

STATUS_COLOR = {
    'POSTED': C['posted'], 'LOCKED': C['locked'],
    'DRAFT': C['draft'], 'SCHEDULED': 'E0F0FF',
}

def fill(h): return PatternFill('solid', fgColor=h)
def font(h='000000', bold=False, size=9): return Font(color=h, bold=bold, size=size, name='Calibri')
def align(h='left', v='top', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def week_of(d):
    if d <= 7: return 1
    if d <= 14: return 2
    if d <= 21: return 3
    if d <= 28: return 4
    return 5

COLUMNS = ['Day #', 'Date', 'Day of Week', 'Week', 'Post Slot', 'Platform',
           'Audience Segment', 'Content Category', 'Hook / Topic',
           'Content Goal', 'Format', 'CTA', 'Monetization',
           'Series Tag', 'Status', 'Notes']
COL_WIDTHS = [6, 11, 12, 7, 10, 13, 16, 16, 60, 13, 18, 20, 13, 18, 11, 40]

# ── POSTS ─────────────────────────────────────────────────────────────────────
# Each tuple: (day, slot, platform, segment, category, hook, goal, format, cta, monetize, series, status, notes)
# Segment rotation enforced: no single segment > 2 consecutive days
# All 4 segments appear in every 7-day week

POSTS = [

    # ══ WEEK 1: RECONNECTION (Days 1–7) ══════════════════════════════════════
    # Segment arc: All → Beginners → Developers → Creators → Business Owners → All → Beginners

    # --- Day 1 (Apr 3) — LOCKED — All segments (comeback story) ---
    (1,'Morning','X/Twitter','All','Personality',
     '"I disappeared. Not because I gave up…" — Comeback thread (5 tweets). CTA: "What did I miss?"',
     'Community','Thread','Reply engagement','N','','LOCKED','Day 1 locked. Already posted.'),
    (1,'Afternoon','X/Twitter','All','Personality',
     '"Most creators disappear because they burned out. I disappeared because I stopped trusting the process."',
     'Reach','Tweet','','N','','LOCKED','Day 1 locked.'),
    (1,'Morning','LinkedIn','All','Personality',
     '"I went silent for months. Not because the business failed…" — Long-form founder comeback story.',
     'Authority','Long-form post','Comments','N','','LOCKED','Day 1 locked.'),
    (1,'Morning','Instagram','All','Personality',
     'Cinematic reel (34s): "What I Built In Silence." Already posted via Graph API.',
     'Reach','Reel','DM CTA','N','','LOCKED','Day 1 locked. Posted.'),
    (1,'Morning','TikTok','All','Personality',
     'Talking head BTS (30s): "I disappeared for months. Here\'s why."',
     'Reach','TikTok video','Follow CTA','N','','LOCKED','Day 1 locked. Posted.'),

    # --- Day 2 (Apr 4) — Beginners: Personality ---
    (2,'Morning','X/Twitter','Beginners','Personality',
     '"The first time I used AI to do something I couldn\'t do myself, I froze. I didn\'t know if it was cheating. Here\'s what I figured out."',
     'Reach','Thread (3 tweets)','Replies','N','','DRAFT',
     'Relatable entry point for beginners. No tech language. Lowers the intimidation barrier.'),
    (2,'Evening','X/Twitter','Beginners','Engagement',
     '"What\'s ONE thing you\'d want AI to do for you if you knew it was possible? Drop it below — I read every reply."',
     'Community','Tweet','Replies (market research)','N','','DRAFT',
     'Surfaces beginner desires. Read replies as demand signal for course content.'),
    (2,'Morning','LinkedIn','Beginners','Value',
     '"The most common AI myth I hear from non-technical people — and why it\'s keeping them stuck."',
     'Authority','Long-form post','Comments','N','','DRAFT','Professional angle on beginner fears.'),
    (2,'Morning','Instagram','Beginners','Value',
     '"5 things beginners get wrong about AI" — Carousel (5 slides). Reassurance format. Dark bg, clean type.',
     'Reach','Carousel','Save + share','N','','DRAFT',''),

    # --- Day 3 (Apr 5) — Developers: Value ---
    (3,'Morning','X/Twitter','Developers','Value',
     '"I replaced 3 hours of manual work with one Claude Code script this week. Here\'s exactly what I built."',
     'Authority','Thread (4 tweets)','Save','N','','DRAFT',
     'Specific + credible. Shows real developer use case without being tutorial-heavy.'),
    (3,'Afternoon','X/Twitter','Developers','Engagement',
     '"What\'s the most repetitive thing in your dev workflow that you haven\'t automated yet? Seriously asking."',
     'Community','Tweet','Replies','N','','DRAFT','Developer audience engagement.'),
    (3,'Morning','TikTok','Developers','Value',
     '"Watch me build a working web scraper in Claude Code in 8 minutes." — Screen recording, speed run.',
     'Reach','TikTok video','Follow for more','N','','DRAFT','Developer-specific content. Speed = hook.'),

    # --- Day 4 (Apr 6) — Creators: Authority ---
    (4,'Morning','X/Twitter','Creators','Authority',
     '"I made a 60-second video this week. Zero filming. Here\'s the 4-tool workflow: Nano Banana → Kling → MiniMax → CapCut."',
     'Authority','Thread (4 tweets)','Save','N','','DRAFT',
     'Creators want the system, not the theory. Show the stack.'),
    (4,'Evening','X/Twitter','Creators','Engagement',
     '"Unpopular opinion: most content creators are overcomplicating video. One AI tool replaced my entire editing setup."',
     'Reach','Tweet','Reply bait','N','','DRAFT','Bold take for creator audience.'),
    (4,'Morning','Instagram','Creators','Value',
     '"My AI video workflow in 60 seconds" — Reel showing Nano Banana → Kling → captions → published post.',
     'Reach','Reel','DM WORKFLOW','N','','DRAFT','Visual proof of the system. No explanation needed.'),

    # --- Day 5 (Apr 7) — Business Owners: Transformational ---
    (5,'Morning','X/Twitter','Business Owners','Transformational',
     '"I stopped doing 14 hours of business tasks manually last month. Here\'s what I automated — and what I did with the time back."',
     'Authority','Thread (4 tweets)','Save','N','','DRAFT',
     'Outcome-led. Business owners respond to time savings, not tool names.'),
    (5,'Morning','LinkedIn','Business Owners','Authority',
     '"Most business owners automate the wrong thing first. Here\'s the order that actually saves real hours, based on what I\'ve tested."',
     'Authority','Long-form post','DM AUTOMATE','N','','DRAFT','LinkedIn audience = business owner decision makers.'),
    (5,'Morning','Instagram','Business Owners','Value',
     '"14 hours back. Here\'s what I automated." — Carousel (5 slides): the 3 automations, time saved each, tools used.',
     'Authority','Carousel','DM AUTOMATE','N','','DRAFT',''),

    # --- Day 6 (Apr 8) — All: Personality ---
    (6,'Morning','X/Twitter','All','Personality',
     '"The hardest part of building something real isn\'t the work. It\'s staying quiet while everyone else chases noise."',
     'Reach','Tweet','Engagement','N','','DRAFT','Builder identity post. Wide appeal.'),
    (6,'Evening','X/Twitter','All','Engagement',
     '"Drop what you\'re building right now. I\'ll reply with one idea to make it 10× better. First 10 replies."',
     'Community','Tweet','Replies','N','','DRAFT','Community engagement. Builds relationship.'),
    (6,'Morning','TikTok','All','Personality',
     '"Behind the scenes of the system I built while I was offline." — BTS of ContentBrain running: skills, calendar, signals.',
     'Reach','TikTok video','Follow CTA','N','','DRAFT','Curiosity + credibility. No explanation of what it is yet.'),

    # --- Day 7 (Apr 9) — Beginners: Value ---
    (7,'Morning','X/Twitter','Beginners','Value',
     '"You don\'t need to understand how AI works to use it. You just need 3 prompts. Here they are."',
     'Reach','Thread (3 tweets)','Save','N','','DRAFT',
     'Lowers the barrier completely. Maximum accessibility.'),
    (7,'Morning','Instagram','Beginners','Freebie',
     '"Your first 3 AI prompts. For total beginners. Save this." — Carousel (4 slides). Clean, simple, dark aesthetic.',
     'Leads','Carousel','Save + share','N','','DRAFT',''),
    (7,'Morning','LinkedIn','Beginners','Authority',
     '"I teach non-technical people to use AI every week. Here\'s the first thing I always tell them."',
     'Authority','Long-form post','Comments','N','','DRAFT','Professional positioning for beginner audience.'),

    # ══ WEEK 2: AUTHORITY + COURSE TEASER (Days 8–14) ═══════════════════════
    # Segment arc: Developers → All → Creators → Business Owners → Beginners → Developers → All

    # --- Day 8 (Apr 10) — Developers: Authority ---
    (8,'Morning','X/Twitter','Developers','Authority',
     '"I audited every AI tool in my stack. Here\'s what survived the cut — and why I dropped the rest."',
     'Authority','Thread (5 tweets)','Save','N','','DRAFT',
     'Data-backed tool analysis. Developers trust opinionated recommendations.'),
    (8,'Afternoon','X/Twitter','Developers','Engagement',
     '"What AI tool surprised you the most this year? One that actually changed your workflow — not just impressed you."',
     'Community','Tweet','Replies','N','','DRAFT',''),
    (8,'Morning','TikTok','Developers','Value',
     '"My actual development workflow with Claude Code — no cuts, no edits." — Screen recording, real project.',
     'Reach','TikTok video','Comment BUILD','N','','DRAFT',''),

    # --- Day 9 (Apr 11) — All: WEEKLY SERIES EP1 (Claude + business analysis) ---
    (9,'Morning','X/Twitter','All','Weekly Series',
     'SERIES EP1: "Things You Didn\'t Know You Could Do With AI #1: Ask Claude to find the gaps in your business in 5 minutes." — Hook + 3 steps.',
     'Authority','Thread (4 tweets)','Save + DM','N','Weekly Series EP1','DRAFT',
     'Tool: Claude. Use case: business gap analysis. Accessible to all 4 segments. Course lead-in: AI Automation Beginner Track M1.'),
    (9,'Morning','Instagram','All','Weekly Series',
     'SERIES EP1: Reel — type your business description into Claude → read the gap analysis output. Show the result.',
     'Authority','Reel (45s)','DM EP1','N','Weekly Series EP1','DRAFT',''),
    (9,'Morning','LinkedIn','All','Value',
     '"I ran an experiment: I described my business to Claude and asked it to find what I was missing. Here\'s what it found."',
     'Authority','Long-form post','DM ANALYSIS','Y','','DRAFT','Subtle course lead-in. Claude capability demo.'),

    # --- Day 10 (Apr 12) — Creators: Value ---
    (10,'Morning','X/Twitter','Creators','Value',
     '"My AI content system explained simply: voice note in → 5 platform-ready posts out. Here\'s every step."',
     'Authority','Thread (4 tweets)','Save','N','','DRAFT',
     'ContentBrain system reveal. Creators want the full workflow.'),
    (10,'Morning','TikTok','Creators','Value',
     '"Voice note in. 5 posts out. Here\'s the full AI content pipeline." — Screen recording of Make.com workflow.',
     'Reach','TikTok video','Follow for more','N','','DRAFT',''),
    (10,'Afternoon','X/Twitter','All','Engagement',
     '"Quick poll: what would you most want AI to handle for you? A) Video creation B) Content writing C) Business automation D) Coding help — reply with your letter."',
     'Community','Tweet','Replies (validation)','N','','DRAFT','Organic validation. Segment signal for course push.'),

    # --- Day 11 (Apr 13) — Business Owners: Authority ---
    (11,'Morning','X/Twitter','Business Owners','Authority',
     '"The automation that gave me 12 hours back last week. Built in an afternoon. Cost: $0 beyond what I already pay."',
     'Authority','Thread (4 tweets)','Save','Y','','DRAFT',
     'Business owners respond to ROI + simplicity. No jargon.'),
    (11,'Morning','LinkedIn','Business Owners','Authority',
     '"The 3 business automations that pay for themselves in week 1. And what to build after that."',
     'Authority','Long-form post','DM AUTOMATE','Y','','DRAFT',''),
    (11,'Morning','Instagram','Business Owners','Value',
     '"The 3 automations every business should build first." — Carousel (5 slides): automation name, time saved, tool used.',
     'Authority','Carousel','DM AUTOMATE','Y','','DRAFT',''),

    # --- Day 12 (Apr 14) — Beginners: Freebie ---
    (12,'Morning','X/Twitter','Beginners','Freebie',
     '"Free: My AI Starter Pack. The 5 prompts I give to anyone who has never used AI before. Repost so it reaches someone who needs it."',
     'Leads','Thread + image','RT + save','Y','','DRAFT',
     'Freebie = algorithm push + goodwill + soft course lead. High repost potential.'),
    (12,'Morning','Telegram','All','Freebie',
     '"AI Beginner Starter Pack — 5 prompts, free. No signup. For anyone who doesn\'t know where to start."',
     'Community','Community post','Download CTA','N','','DRAFT','Telegram community activation.'),
    (12,'Morning','Instagram','Beginners','Freebie',
     '"5 AI starter prompts. Free. Save this." — Carousel (6 slides). One prompt per slide, what it does, try it now.',
     'Reach','Carousel','Save + share','N','','DRAFT',''),

    # --- Day 13 (Apr 15) — Developers: Transformational ---
    (13,'Morning','X/Twitter','Developers','Transformational',
     '"What changed when I stopped writing code from scratch: I ship 3× more, my clients get faster results, and I focus on architecture instead of syntax."',
     'Authority','Thread (4 tweets)','Save','N','','DRAFT','Developer identity shift.'),
    (13,'Morning','TikTok','Developers','Value',
     '"Before vs after Claude Code. Same project. Real-time comparison." — Split screen screen recording.',
     'Reach','TikTok video','Comment BEFORE','N','','DRAFT',''),

    # --- Day 14 (Apr 16) — All: Personality + segment self-selection ---
    (14,'Morning','X/Twitter','All','Personality',
     '"Hot take: most people aren\'t behind on AI. They\'re paralyzed by too many options. The fix is simpler than you think."',
     'Reach','Tweet','Replies','N','','DRAFT','Wide appeal. Pre-validation energy.'),
    (14,'Afternoon','X/Twitter','All','Engagement',
     '"Which of these is you right now? A) Total AI beginner B) Learning to automate C) Building with code D) Using AI in my business — reply with your letter."',
     'Community','Tweet','Replies (segmentation)','N','','DRAFT',
     'Audience segmentation post. Read replies to calibrate Week 3 content.'),
    (14,'Morning','Instagram','All','Personality',
     '"4 types of people who need AI right now." — Carousel. One slide per segment. Which one are you? CTA: DM your letter.',
     'Leads','Carousel','DM A/B/C/D','Y','','DRAFT','Soft lead-in to course announcement.'),

    # ══ WEEK 3: ANNOUNCEMENT + VALIDATION (Days 15–21) ══════════════════════
    # Segment arc: All (validation) → Business/Creators → All (announce) → Beginners → Creators → Developers → All

    # --- Day 15 (Apr 17) — All: VALIDATION POSTS ---
    (15,'Morning','X/Twitter','Beginners','Lead Gen',
     'VALIDATION: "If someone ran live AI classes twice a week — no coding required, just workflows that save time — would you join? What would make you say yes immediately?"',
     'Leads','Tweet','Reply (validation)','Y','Validation Post','DRAFT',
     'Publish exactly Day 15. Read every reply within 2 hours. Measure real demand.'),
    (15,'Morning','LinkedIn','Developers','Lead Gen',
     'VALIDATION: "Thinking about running a live AI engineering cohort — Claude Code, multi-agent systems, full-stack AI apps. What\'s the one thing you\'ve never seen taught properly in an AI course?"',
     'Leads','Post + poll','Comments','Y','Validation Post','DRAFT','LinkedIn dev segment.'),
    (15,'Morning','Instagram','Creators','Lead Gen',
     'VALIDATION: "Would you join a live AI video creation course? No camera. No editing. Just the workflow." — Story poll + reel.',
     'Leads','Story poll + Reel','Vote + DM VIDEOCLASS','Y','Validation Post','DRAFT',''),
    (15,'Morning','TikTok','Business Owners','Lead Gen',
     'VALIDATION: "I\'m thinking about running a live AI course for business owners. Would you join if I showed you how to save 10+ hours a week? React if yes." — 30s talking head.',
     'Leads','TikTok video','React/comment','Y','Validation Post','DRAFT',''),

    # --- Day 16 (Apr 18) — Business Owners + Creators: Money Proof ---
    (16,'Morning','X/Twitter','All','Money Proof',
     'INCOME BREAKDOWN: "5 ways people are making real money with AI tools in 2026. With numbers and sources." — Thread.',
     'Authority','Thread (6 tweets)','Save + share','Y','Money Proof EP1','DRAFT',
     'Streams: freelancing $2-8K/mo, Claude Code $500-10K/mo, automation services $800+$200/mo, AI video $200-5K/deal, consulting $5-50K/mo. Source every claim.'),
    (16,'Morning','LinkedIn','Business Owners','Money Proof',
     '"What AI consultants are charging in 2026. And how non-technical beginners are starting at $500/month." — Income data post.',
     'Authority','Long-form post','DM INCOME','Y','Money Proof EP1','DRAFT',''),
    (16,'Morning','Instagram','Creators','Money Proof',
     '"How people are making money with AI — 5 real income streams." — Carousel (6 slides). Numbers sourced and labeled.',
     'Authority','Carousel','Save + share','Y','Money Proof EP1','DRAFT',''),

    # --- Day 17 (Apr 19) — All: Course Announcement ---
    (17,'Morning','X/Twitter','All','Course Launch',
     '"Something I\'ve been building for 30 days. A live AI course. 3 topics. 4 audience tracks. $15–$50. Starting May 3. Here\'s everything inside." — Announcement thread.',
     'Leads','Thread (5 tweets)','DM COURSE','Y','Course Announcement','DRAFT',
     'First hard announcement. Clear, specific, confident. No desperation.'),
    (17,'Morning','LinkedIn','All','Course Launch',
     '"I\'m launching an AI course in 16 days. Live classes twice a week. 4 audience tracks. Priced for accessibility. Here\'s why I built it."',
     'Leads','Long-form post','DM COURSE','Y','Course Announcement','DRAFT',''),
    (17,'Morning','Telegram','All','Course Launch',
     '"IT\'S LIVE. Quivira AI Course announced. 3 topics. 4 tracks. 30 days. Link inside. Founding prices end May 2."',
     'Community','Community post + pin','Click link','Y','Course Announcement','DRAFT','Pin this message.'),
    (17,'Morning','Instagram','All','Course Launch',
     '"The course is coming." — Cinematic reveal reel (30s). Dark. Show the framework, the tools, the stack. No price yet.',
     'Reach','Reel','DM COURSE','Y','Course Announcement','DRAFT',''),

    # --- Day 18 (Apr 20) — Beginners: Value (course content preview) ---
    (18,'Morning','X/Twitter','Beginners','Value',
     '"How to set up your first AI automation with zero prior experience. The exact steps. Takes under an hour." — Tutorial thread.',
     'Authority','Thread (5 tweets)','Save + DM','Y','','DRAFT',
     'Beginner course content preview. Shows what they get inside.'),
    (18,'Morning','TikTok','Beginners','Value',
     '"Zero experience → first automation in under an hour. Watch me do it." — Speed run screen recording.',
     'Reach','TikTok video','Comment AUTOMATE','Y','','DRAFT',''),
    (18,'Afternoon','X/Twitter','All','Lead Gen',
     '"Early access to the Quivira AI Course is open. $15 / $30 / $50 depending on tier. DM me COURSE for the link."',
     'Leads','Tweet','DM COURSE','Y','','DRAFT',''),

    # --- Day 19 (Apr 21) — Creators: WEEKLY SERIES EP2 (Kling AI) ---
    (19,'Morning','X/Twitter','Creators','Weekly Series',
     'SERIES EP2: "Things You Didn\'t Know You Could Do With AI #2: Generate a cinematic video from a single text prompt using Kling. No camera. No editor." — Hook + steps.',
     'Authority','Thread (4 tweets)','Save + DM','Y','Weekly Series EP2','DRAFT',
     'Tool: Kling AI. Use case: social media video creation. Lead-in: AI Video Creation course.'),
    (19,'Morning','Instagram','Creators','Weekly Series',
     'SERIES EP2: Reel — text prompt → 5s cinematic video → published post. Real time. 45s.',
     'Reach','Reel','DM KLING','Y','Weekly Series EP2','DRAFT',''),
    (19,'Morning','LinkedIn','Creators','Value',
     '"Why every content creator needs an AI video workflow in 2026. And why most still don\'t have one."',
     'Authority','Long-form post','DM VIDEO','Y','','DRAFT',''),

    # --- Day 20 (Apr 22) — Developers: Authority ---
    (20,'Morning','X/Twitter','Developers','Authority',
     '"I built a working full-stack web app in 4 hours using Claude Code and zero prior framework knowledge. Here\'s what happened." — Build thread.',
     'Authority','Thread (5 tweets)','Save','Y','','DRAFT','Developer credibility post.'),
    (20,'Morning','TikTok','Developers','Value',
     '"4 hours. Full-stack app. Claude Code. Watch the build." — Screen recording timelapse.',
     'Reach','TikTok video','Comment BUILD','Y','','DRAFT',''),

    # --- Day 21 (Apr 23) — All: Lead Gen + pricing reveal ---
    (21,'Morning','X/Twitter','All','Lead Gen',
     '"The Quivira AI Course has 3 tiers:\n$15 — self-paced (pre-recorded)\n$30 — live classes + recordings\n$50 — full access + direct DM to me + 2 exclusive bonuses\nWhich one is you? DM me the price."',
     'Leads','Tweet','DM tier','Y','','DRAFT','Tier reveal. Simple, clear, no fluff.'),
    (21,'Morning','LinkedIn','Business Owners','Lead Gen',
     '"I\'m running this course at founding prices because I want proof it works. Here\'s exactly what each tier gets you." — Pricing breakdown post.',
     'Leads','Long-form post','DM PRICING','Y','','DRAFT',''),
    (21,'Morning','Telegram','All','Lead Gen',
     '"Full Access (all 3 topics + direct DM access to me) is capped at 20. Seats are going. Here\'s the link."',
     'Community','Community post','Click link','Y','','DRAFT',''),

    # ══ WEEK 4: ACTIVE PROMOTION + SOCIAL PROOF (Days 22–28) ══════════════
    # Segment arc: Beginners → All → Developers → Creators → Business Owners → All → All

    # --- Day 22 (Apr 24) — Beginners: Money Proof ---
    (22,'Morning','X/Twitter','Beginners','Money Proof',
     'CASE STUDY: "A freelance writer with zero coding experience raised her rates and works half the hours. She uses Claude for research. Here\'s the exact method." — Breakdown thread.',
     'Authority','Thread (4 tweets)','Save + share','Y','Money Proof EP2','DRAFT',
     'Label as composite example based on industry data. $2,400 → $3,600/mo same hours.'),
    (22,'Morning','Instagram','Beginners','Money Proof',
     '"From $2,400 to $3,600/month. No coding. No extra hours. Just Claude." — Carousel (5 slides).',
     'Authority','Carousel','Save + DM','Y','Money Proof EP2','DRAFT',''),
    (22,'Afternoon','X/Twitter','All','Course Launch',
     '"9 days left. Quivira AI Course. $15–$50. Live classes start May 3. DM me COURSE."',
     'Leads','Tweet','DM COURSE','Y','','DRAFT','Countdown urgency.'),

    # --- Day 23 (Apr 25) — All: Social Proof ---
    (23,'Morning','X/Twitter','All','Course Launch',
     '"Students are inside the course. Here\'s what they\'re saying after seeing the full curriculum." — Quote-style thread.',
     'Leads','Thread (3 tweets)','DM COURSE','Y','Social Proof','DRAFT',
     'Use validation poll responses or early student reactions as social proof.'),
    (23,'Morning','Instagram','All','Course Launch',
     '"What\'s inside the Quivira AI Course." — Carousel walkthrough (8 slides): topics, tracks, live sessions, bonuses, pricing.',
     'Leads','Carousel','DM COURSE','Y','','DRAFT',''),
    (23,'Morning','TikTok','All','Course Launch',
     '"I\'m running a live AI course for 4 weeks. Here\'s what Week 1 looks like inside." — BTS footage.',
     'Leads','TikTok video','Link in bio','Y','','DRAFT',''),

    # --- Day 24 (Apr 26) — Developers: WEEKLY SERIES EP3 + Freebie ---
    (24,'Morning','X/Twitter','Developers','Weekly Series',
     'SERIES EP3: "Things You Didn\'t Know You Could Do With AI #3: Query a database in plain English using Claude + Supabase. No SQL. No developer." + FREE: Module 1 pre-recorded for 24hrs.',
     'Leads','Thread (4 tweets)','DM MODULE1','Y','Weekly Series EP3','DRAFT',
     'Tool: Supabase + Claude. Double value: series + freebie. Drives course interest for all segments.'),
    (24,'Morning','Telegram','All','Freebie',
     '"FREE for 24 hours: Module 1 of the AI Automation course. Pre-recorded. Zero payment required. Link below."',
     'Leads','Community post','Click link','Y','Weekly Series EP3','DRAFT',''),
    (24,'Morning','Instagram','Developers','Weekly Series',
     'SERIES EP3: Reel — type English question → Claude writes SQL → Supabase returns answer. 45s.',
     'Authority','Reel','DM MODULE1','Y','Weekly Series EP3','DRAFT',''),

    # --- Day 25 (Apr 27) — Creators: Social Proof + BTS ---
    (25,'Morning','X/Twitter','Creators','Personality',
     '"Week 1 inside the AI Video Creation track. Here\'s what students produced in their first 90 minutes. Without filming anything." — Student output thread.',
     'Authority','Thread (3 tweets)','DM COURSE','Y','Social Proof','DRAFT',''),
    (25,'Morning','TikTok','Creators','Course Launch',
     '"What students built in Week 1 of the AI video course — no camera, no studio." — Output montage or BTS.',
     'Leads','TikTok video','Link in bio','Y','','DRAFT',''),

    # --- Day 26 (Apr 28) — Business Owners: Authority ---
    (26,'Morning','X/Twitter','Business Owners','Authority',
     '"The 5 AI skills that will define what you can charge in 2027. And which one to start with if you\'re non-technical." — Thread.',
     'Authority','Thread (5 tweets)','Save','N','','DRAFT','Authority piece. Broad appeal. Not promotional.'),
    (26,'Morning','LinkedIn','Business Owners','Authority',
     '"The 5 AI skills companies will pay premium for in 2027. How to build all of them — even without a technical background."',
     'Authority','Long-form post','DM for more','Y','','DRAFT',''),

    # --- Day 27 (Apr 29) — All: Urgency ---
    (27,'Morning','X/Twitter','All','Course Launch',
     '"3 days left to join Quivira AI Course at founding prices. After May 2, prices go up for Cohort 2."',
     'Leads','Tweet','DM COURSE','Y','','DRAFT','Genuine urgency — Cohort 1 pricing is locked.'),
    (27,'Afternoon','X/Twitter','All','Engagement',
     '"What would stop you from joining a live AI course right now? I\'ll answer every reply honestly."',
     'Community','Tweet','Replies (objections)','Y','','DRAFT','Objection handling. Read replies for live rebuttals.'),
    (27,'Morning','LinkedIn','All','Course Launch',
     '"3 days left. If you\'ve been watching from the sidelines — this is the moment." — Final LinkedIn pitch.',
     'Leads','Post','DM COURSE','Y','','DRAFT',''),
    (27,'Morning','Telegram','All','Course Launch',
     '"3 days. Founding prices close May 2. Full Access almost full. Drop your question here — I\'ll answer personally."',
     'Community','Community post','Replies','Y','','DRAFT',''),

    # --- Day 28 (Apr 30) — All: Transformational (student wins) ---
    (28,'Morning','X/Twitter','Beginners','Transformational',
     '"Student win from inside the course: automated their entire client onboarding process in under 90 minutes. Zero prior AI experience before they enrolled." — Win showcase.',
     'Leads','Tweet','DM COURSE','Y','Social Proof','DRAFT',''),
    (28,'Morning','Instagram','All','Transformational',
     '"What students are building inside the Quivira AI Course — real Week 1 results." — Reel or carousel of student outputs.',
     'Leads','Reel','DM COURSE','Y','Social Proof','DRAFT',''),

    # ══ DAYS 29–30: FINAL PUSH ════════════════════════════════════════════════

    # --- Day 29 (May 1) — All: FOMO ---
    (29,'Morning','X/Twitter','All','Course Launch',
     '"Last 24 hours. Quivira AI Course Cohort 1 closes tonight. After this, prices go up and the ContentBrain bonus is gone."',
     'Leads','Thread (3 tweets)','DM COURSE','Y','','DRAFT','Final push. Honest, not desperate.'),
    (29,'Afternoon','X/Twitter','All','Lead Gen',
     '"Full Access (all 3 topics + direct DM to me) — X seats remaining. Once we close, we close."',
     'Leads','Tweet','DM FULLACCESS','Y','','DRAFT','Real scarcity drives action.'),
    (29,'Morning','LinkedIn','All','Course Launch',
     '"Today is the last day to join Cohort 1 at founding prices. What you get for $50 is worth 10× that."',
     'Leads','Post','DM COURSE','Y','','DRAFT',''),
    (29,'Morning','Telegram','All','Course Launch',
     '"LAST DAY. Cohort 1 closes tonight. Link inside. Who\'s still on the fence — reply here."',
     'Community','Community post + pin','Click link','Y','','DRAFT','Pin this.'),
    (29,'Morning','Instagram','All','Course Launch',
     '"Last 24 hours." — Text animation on dark background. Simple. Countdown energy.',
     'Leads','Reel (15s)','Link in bio','Y','','DRAFT',''),

    # --- Day 30 (May 2) — All: Final Day + WEEKLY SERIES EP4 (Remotion) ---
    (30,'Morning','X/Twitter','Developers','Weekly Series',
     'SERIES EP4: "Things You Didn\'t Know You Could Do With AI #4: Generate 30 videos from a CSV file using Remotion. One command. No manual editing." — Thread.',
     'Authority','Thread (4 tweets)','DM REMOTION','Y','Weekly Series EP4','DRAFT',
     'Tool: Remotion. Lead-in: Coding with AI Developer Track. Final series episode.'),
    (30,'Afternoon','X/Twitter','All','Course Launch',
     '"Today is the last day. Here\'s everything you get inside the Quivira AI Course." — Full summary tweet.',
     'Leads','Thread (5 tweets)','DM COURSE','Y','','DRAFT','Final pitch. Complete. Confident.'),
    (30,'Morning','LinkedIn','All','Personality',
     '"Today Cohort 1 closes. Thank you to everyone who joined. Here\'s what we\'re going to build together over the next 30 days." — Gratitude + vision.',
     'Community','Post','DM COURSE','Y','','DRAFT',''),
    (30,'Morning','Instagram','Creators','Weekly Series',
     'SERIES EP4: Reel — CSV → 30 videos generated automatically. Terminal render. Folder of video files. 45s.',
     'Authority','Reel','DM REMOTION','Y','Weekly Series EP4','DRAFT',''),
    (30,'Morning','TikTok','All','Course Launch',
     '"The Quivira AI Course starts tomorrow. Here\'s what Cohort 1 looks like inside." — Final hype video.',
     'Leads','TikTok video','Link in bio','Y','','DRAFT',''),
]


# ── BUILD CALENDAR ────────────────────────────────────────────────────────────

def sc(cell, value, bg=C['white'], fg=C['gray'], bold=False,
       h='left', v='top', wrap=True):
    cell.value = value
    cell.fill = fill(bg)
    cell.font = font(fg, bold)
    cell.alignment = align(h, v, wrap)
    cell.border = border()

def sh(cell, text, bg=C['dark'], fg=C['white'], bold=True):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = font(fg, bold)
    cell.alignment = align('center', 'center', False)
    cell.border = border()


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = '30-Day Launch Calendar'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')
    t = ws['A1']
    t.value = ('QUIVIRA AI COURSE — 30-Day Launch Calendar  |  April 3–May 2, 2026  '
               '|  Day 1 = LOCKED  |  v2: Balanced across all 4 audience segments')
    t.fill = fill(C['dark'])
    t.font = font(C['white'], True, 11)
    t.alignment = align('center', 'center', False)
    ws.row_dimensions[1].height = 22

    # Column headers + widths
    ws.row_dimensions[2].height = 18
    for ci, (col, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        sh(ws.cell(2, ci), col)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    row = 3
    for post in POSTS:
        (day_num, slot, platform, segment, category,
         hook, goal, fmt, cta, monetize, series, status, notes) = post

        day_date = START + timedelta(days=day_num - 1)
        wk = week_of(day_num)
        wk_bg = WEEK_BG.get(wk, C['white'])
        plat_col = PLATFORM_COLOR.get(platform, C['dark'])
        cat_bg = CATEGORY_COLOR.get(category, C['white'])
        st_bg = STATUS_COLOR.get(status, C['white'])

        values = [
            day_num,
            day_date.strftime('%b %d'),
            day_date.strftime('%A'),
            f'Week {wk}',
            slot, platform, segment, category, hook,
            goal, fmt, cta, monetize, series, status, notes,
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row, ci)
            col_name = COLUMNS[ci - 1]

            if col_name == 'Platform':
                sc(cell, val, plat_col, C['white'], bold=True, h='center')
            elif col_name == 'Content Category':
                sc(cell, val, cat_bg, C['gray'], h='center')
            elif col_name == 'Status':
                sc(cell, val, st_bg, C['dark'], bold=(status == 'LOCKED'), h='center')
            elif col_name == 'Monetization':
                mb = 'E5FFE5' if val == 'Y' else C['white']
                sc(cell, val, mb, C['gray'], bold=(val == 'Y'), h='center')
            elif col_name in ('Hook / Topic', 'Notes'):
                sc(cell, val, C['white'], C['gray'], h='left', v='top', wrap=True)
            elif col_name == 'Series Tag':
                sb = 'CCF5FF' if val else C['white']
                sc(cell, val, sb, C['dark'], h='center')
            elif col_name == 'Day #':
                db = C['locked'] if status == 'LOCKED' else wk_bg
                sc(cell, val, db, C['dark'], bold=True, h='center')
            elif col_name in ('Date', 'Day of Week', 'Week'):
                sc(cell, val, wk_bg, C['gray'], h='center')
            else:
                sc(cell, val, C['white'], C['gray'], h='left')

        ws.row_dimensions[row].height = 42
        row += 1

    # Segment balance summary
    ws.append([])
    ws.append([])
    sr = row + 2
    ws.cell(sr, 1).value = 'SEGMENT DISTRIBUTION SUMMARY (Days 2–30)'
    ws.cell(sr, 1).font = font(C['dark'], True, 10)

    seg_counts = {}
    for post in POSTS:
        if post[0] == 1:
            continue
        seg = post[3]
        seg_counts[seg] = seg_counts.get(seg, 0) + 1
    total = sum(seg_counts.values())

    for i, (seg, count) in enumerate(sorted(seg_counts.items())):
        r = sr + 1 + i
        ws.cell(r, 1).value = seg
        ws.cell(r, 1).font = font(C['dark'], True)
        ws.cell(r, 1).border = border()
        ws.cell(r, 2).value = f'{count} posts ({round(count/total*100)}%)'
        ws.cell(r, 2).font = font(C['gray'])

    # Legend
    lr = sr + len(seg_counts) + 2
    ws.cell(lr, 1).value = 'LEGEND'
    ws.cell(lr, 1).font = font(C['dark'], True, 10)
    for i, (s, bg, desc) in enumerate([
        ('LOCKED', C['locked'], 'Day 1 — already posted, do not modify'),
        ('DRAFT',  C['draft'],  'Written, not yet scheduled'),
        ('SCHEDULED', 'E0F0FF', 'Queued in Typefully / Buffer'),
        ('POSTED', C['posted'], 'Confirmed published'),
    ]):
        r = lr + 1 + i
        ws.cell(r, 1).value = s
        ws.cell(r, 1).fill = fill(bg)
        ws.cell(r, 1).font = font(C['dark'], True)
        ws.cell(r, 1).border = border()
        ws.cell(r, 2).value = desc
        ws.cell(r, 2).font = font(C['gray'])

    wb.save(OUT_PATH)
    print(f'Calendar saved: {OUT_PATH}')

    # Print segment summary
    print('\nSegment distribution (Days 2–30):')
    for seg, count in sorted(seg_counts.items()):
        print(f'  {seg}: {count} posts ({round(count/total*100)}%)')


if __name__ == '__main__':
    build()
